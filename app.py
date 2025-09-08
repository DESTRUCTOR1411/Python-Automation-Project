import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # cell = sheet.cell(1,1)
    # print(sheet.max_row)

    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row,3)
        # print(type(cell.value))
        if cell.value is None:   # skip blank cells
            continue
        raw_value = str(cell.value).replace("$", "").replace(",", "").strip()
        corrected_value= float(raw_value) * 0.5
        corrected_price_cell= sheet.cell(row,4)
        corrected_price_cell.value= corrected_value


    # we will take the reference of values
    values= Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

    # lets create the chart -- barchart
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')


    wb.save(filename)
